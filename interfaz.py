import sys
import os
import csv
import json
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                            QLabel, QLineEdit, QPushButton, QMessageBox, QGridLayout,
                            QFileDialog, QTabWidget, QProgressBar, QHBoxLayout, QFrame,
                            QScrollArea, QTableWidget, QTableWidgetItem, QHeaderView,
                            QAbstractItemView)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QIcon
from calcular_hash import HashLogica
from translations import translations

class HashCalculatorThread(QThread):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    progress = pyqtSignal(int)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.hash_logic = HashLogica()

    def run(self):
        try:
            result = self.hash_logic.calculate_file_hash(
                self.file_path, self.progress.emit)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))

class DropLineEdit(QLineEdit):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            self.setText(files[0])

class MultiDropLineEdit(QLineEdit):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.files = []

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            self.files = files
            self.setText(", ".join([os.path.basename(f) for f in files]))

class HashGUI(QMainWindow):

    def __init__(self):
        super().__init__()
        self.hash_logic = HashLogica()
        self.language = 'es'
        self.dark_mode = True
        self.text_history = []
        self.file_history = []
        self.setup_ui()

    def tr(self, key):
        return translations[self.language].get(key, key)

    def setup_ui(self):
        self.setWindowTitle(self.tr('window_title'))
        self.setFixedSize(1100, 550) # Tamaño fijo
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowMaximizeButtonHint) # Desactivar maximizar
        
        icon_path = os.path.join(os.path.dirname(__file__), 'icono.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 5, 10, 10)
        main_layout.setSpacing(0)

        # Tab Widget
        self.tab_widget = QTabWidget()
        
        # Widget para los botones en la esquina de las pestañas
        corner_widget = QWidget()
        corner_layout = QHBoxLayout(corner_widget)
        corner_layout.setContentsMargins(0, 0, 5, 0)
        corner_layout.setSpacing(5)
        
        self.language_btn = QPushButton(self.tr('english') if self.language == 'es' else self.tr('spanish'))
        self.language_btn.setFixedWidth(100)
        self.language_btn.clicked.connect(self.toggle_language)
        
        self.toggle_mode_btn = QPushButton("🌙" if not self.dark_mode else "🌞")
        self.toggle_mode_btn.setObjectName("theme_toggle_btn")
        self.toggle_mode_btn.setFixedSize(40, 40)
        self.toggle_mode_btn.clicked.connect(self.toggle_dark_mode)
        
        corner_layout.addWidget(self.language_btn)
        corner_layout.addWidget(self.toggle_mode_btn)
        
        self.tab_widget.setCornerWidget(corner_widget, Qt.Corner.TopRightCorner)
        main_layout.addWidget(self.tab_widget)

        # Setup de pestañas
        self.setup_text_tab()
        self.setup_file_tab()
        self.setup_verify_tab()
        self.setup_compare_folders_tab()
        self.setup_batch_hash_tab()
        self.setup_duplicate_finder_tab()
        self.setup_pro_reports_tab()

        self.apply_styles()

    def setup_text_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        input_layout = QHBoxLayout()
        self.text_input = QLineEdit()
        self.text_input.setPlaceholderText(self.tr('text_input_placeholder'))
        calc_button = QPushButton(self.tr('calculate'))
        calc_button.clicked.connect(self.calculate_text_hash)
        
        input_layout.addWidget(QLabel(self.tr('text_input_label')))
        input_layout.addWidget(self.text_input)
        input_layout.addWidget(calc_button)
        layout.addLayout(input_layout)

        # Resultados
        self.text_results = {}
        results_layout = QGridLayout()
        
        algorithms = ['md5', 'sha1', 'sha224', 'sha256', 'sha384', 'sha512', 'blake2b', 'blake2s', 'sha3_256', 'sha3_512', 'crc32']
        for i, algo in enumerate(algorithms):
            row = i % 6
            col_offset = (i // 6) * 3
            
            label = QLabel(f"<b>{algo.upper()}:</b>")
            result = QLineEdit()
            result.setReadOnly(True)
            copy_btn = QPushButton(self.tr('copy'))
            copy_btn.setFixedWidth(70)
            copy_btn.clicked.connect(lambda checked, r=result: self.copy_to_clipboard(r.text()))
            
            results_layout.addWidget(label, row, col_offset)
            results_layout.addWidget(result, row, col_offset + 1)
            results_layout.addWidget(copy_btn, row, col_offset + 2)
            self.text_results[algo] = result
            
        layout.addLayout(results_layout)

        # Historial
        history_layout = QHBoxLayout()
        self.text_history_label = QLabel(self.tr('hash_history_text'))
        self.text_history_last = QLabel("")
        export_btn = QPushButton(self.tr('export_text_history'))
        export_btn.clicked.connect(self.export_text_history)
        
        history_layout.addWidget(self.text_history_label)
        history_layout.addWidget(self.text_history_last)
        history_layout.addStretch()
        history_layout.addWidget(export_btn)
        layout.addLayout(history_layout)

        self.tab_widget.addTab(tab, self.tr('text_tab'))

    def setup_file_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        input_layout = QHBoxLayout()
        self.file_input = DropLineEdit()
        self.file_input.setPlaceholderText(self.tr('select_file'))
        browse_btn = QPushButton(self.tr('browse'))
        browse_btn.clicked.connect(self.browse_file)
        calc_btn = QPushButton(self.tr('calculate'))
        calc_btn.clicked.connect(self.calculate_file_hash)
        
        input_layout.addWidget(QLabel(self.tr('file_input_label')))
        input_layout.addWidget(self.file_input)
        input_layout.addWidget(browse_btn)
        input_layout.addWidget(calc_btn)
        layout.addLayout(input_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # Resultados
        self.file_results = {}
        results_layout = QGridLayout()
        
        algorithms = ['md5', 'sha1', 'sha224', 'sha256', 'sha384', 'sha512', 'blake2b', 'blake2s', 'sha3_256', 'sha3_512', 'crc32']
        for i, algo in enumerate(algorithms):
            row = i % 6
            col_offset = (i // 6) * 3
            
            label = QLabel(f"<b>{algo.upper()}:</b>")
            result = QLineEdit()
            result.setReadOnly(True)
            copy_btn = QPushButton(self.tr('copy'))
            copy_btn.setFixedWidth(70)
            copy_btn.clicked.connect(lambda checked, r=result: self.copy_to_clipboard(r.text()))
            
            results_layout.addWidget(label, row, col_offset)
            results_layout.addWidget(result, row, col_offset + 1)
            results_layout.addWidget(copy_btn, row, col_offset + 2)
            self.file_results[algo] = result

        layout.addLayout(results_layout)

        # Historial
        history_layout = QHBoxLayout()
        self.file_history_label = QLabel(self.tr('hash_history_file'))
        self.file_history_last = QLabel("")
        export_btn = QPushButton(self.tr('export_file_history'))
        export_btn.clicked.connect(self.export_file_history)
        
        history_layout.addWidget(self.file_history_label)
        history_layout.addWidget(self.file_history_last)
        history_layout.addStretch()
        history_layout.addWidget(export_btn)
        layout.addLayout(history_layout)

        self.tab_widget.addTab(tab, self.tr('file_tab'))

    def setup_verify_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Modo integridad
        integrity_group = QFrame()
        integrity_group.setFrameShape(QFrame.Shape.StyledPanel)
        integrity_layout = QVBoxLayout(integrity_group)
        
        integrity_layout.addWidget(QLabel(f"<b>{self.tr('btn_verify_file')}</b>"))
        
        file1_layout = QHBoxLayout()
        self.verify_file_input = DropLineEdit()
        browse_btn = QPushButton(self.tr('browse'))
        browse_btn.clicked.connect(lambda: self.browse_verify_file(0))
        file1_layout.addWidget(QLabel(self.tr('verify_file_label')))
        file1_layout.addWidget(self.verify_file_input)
        file1_layout.addWidget(browse_btn)
        integrity_layout.addLayout(file1_layout)
        
        hash_input_layout = QHBoxLayout()
        self.verify_hash_input = QLineEdit()
        self.verify_hash_input.setPlaceholderText(self.tr('expected_hash_placeholder'))
        verify_btn = QPushButton(self.tr('btn_verify_file'))
        verify_btn.clicked.connect(self.verify_hash)
        hash_input_layout.addWidget(QLabel(self.tr('expected_hash_label')))
        hash_input_layout.addWidget(self.verify_hash_input)
        hash_input_layout.addWidget(verify_btn)
        integrity_layout.addLayout(hash_input_layout)
        
        layout.addWidget(integrity_group)
        
        # Separador
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)
        
        # Modo Comparar archivos
        compare_group = QFrame()
        compare_group.setFrameShape(QFrame.Shape.StyledPanel)
        compare_layout = QVBoxLayout(compare_group)
        
        compare_layout.addWidget(QLabel(f"<b>{self.tr('btn_compare_files')}</b>"))
        
        f1_layout = QHBoxLayout()
        self.compare_file1_input = DropLineEdit()
        browse1_btn = QPushButton(self.tr('browse'))
        browse1_btn.clicked.connect(lambda: self.browse_verify_file(1))
        f1_layout.addWidget(QLabel(self.tr('compare_file1')))
        f1_layout.addWidget(self.compare_file1_input)
        f1_layout.addWidget(browse1_btn)
        compare_layout.addLayout(f1_layout)
        
        f2_layout = QHBoxLayout()
        self.compare_file2_input = DropLineEdit()
        browse2_btn = QPushButton(self.tr('browse'))
        browse2_btn.clicked.connect(lambda: self.browse_verify_file(2))
        f2_layout.addWidget(QLabel(self.tr('compare_file2')))
        f2_layout.addWidget(self.compare_file2_input)
        f2_layout.addWidget(browse2_btn)
        compare_layout.addLayout(f2_layout)
        
        compare_btn = QPushButton(self.tr('btn_compare_files'))
        compare_btn.clicked.connect(self.compare_files)
        compare_layout.addWidget(compare_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(compare_group)
        
        # Estado
        self.verify_status_label = QLabel("")
        self.verify_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.verify_details_label = QLabel("")
        self.verify_details_label.setWordWrap(True)
        layout.addWidget(self.verify_status_label)
        layout.addWidget(self.verify_details_label)
        
        layout.addStretch()
        self.tab_widget.addTab(tab, self.tr('verify_tab'))

    def setup_compare_folders_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        f1_layout = QHBoxLayout()
        self.folder1_input = QLineEdit()
        self.folder1_input.setReadOnly(True)
        browse1_btn = QPushButton(self.tr('browse'))
        browse1_btn.clicked.connect(lambda: self.browse_folder(1))
        f1_layout.addWidget(QLabel(self.tr('compare_folder1')))
        f1_layout.addWidget(self.folder1_input)
        f1_layout.addWidget(browse1_btn)
        layout.addLayout(f1_layout)
        
        f2_layout = QHBoxLayout()
        self.folder2_input = QLineEdit()
        self.folder2_input.setReadOnly(True)
        browse2_btn = QPushButton(self.tr('browse'))
        browse2_btn.clicked.connect(lambda: self.browse_folder(2))
        f2_layout.addWidget(QLabel(self.tr('compare_folder2')))
        f2_layout.addWidget(self.folder2_input)
        f2_layout.addWidget(browse2_btn)
        layout.addLayout(f2_layout)
        
        compare_btn = QPushButton(self.tr('btn_compare_folders'))
        compare_btn.clicked.connect(self.compare_folders)
        layout.addWidget(compare_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        self.folder_compare_results = QLabel("")
        self.folder_compare_results.setWordWrap(True)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.folder_compare_results)
        layout.addWidget(scroll)
        
        self.tab_widget.addTab(tab, self.tr('compare_folders'))

    def setup_batch_hash_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        controls = QHBoxLayout()
        add_btn = QPushButton(self.tr('add_files'))
        add_btn.clicked.connect(self.add_files_to_batch)
        clear_btn = QPushButton(self.tr('clear'))
        clear_btn.clicked.connect(self.clear_batch)
        process_btn = QPushButton(self.tr('calculate'))
        process_btn.clicked.connect(self.process_batch)
        export_btn = QPushButton(self.tr('export_batch_results'))
        export_btn.clicked.connect(self.export_batch_excel)
        
        controls.addWidget(add_btn)
        controls.addWidget(clear_btn)
        controls.addStretch()
        controls.addWidget(process_btn)
        controls.addWidget(export_btn)
        layout.addLayout(controls)
        
        self.batch_table = QTableWidget(0, 4)
        self.batch_table.setHorizontalHeaderLabels([
            self.tr('col_file'), self.tr('col_size'), self.tr('col_md5'), self.tr('col_sha256')
        ])
        self.batch_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.batch_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.batch_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.batch_table.setAcceptDrops(True)
        # Habilitar DnD en la tabla
        self.batch_table.dragEnterEvent = self.table_dragEnterEvent
        self.batch_table.dropEvent = self.table_dropEvent
        
        layout.addWidget(self.batch_table)
        
        self.tab_widget.addTab(tab, self.tr('batch_tab'))

    def table_dragEnterEvent(self, event):
        if event.mimeData().hasUrls(): event.acceptProposedAction()

    def table_dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls() if os.path.isfile(u.toLocalFile())]
        self.add_files_to_table(files)

    def add_files_to_batch(self):
        files, _ = QFileDialog.getOpenFileNames(self, self.tr('add_files'))
        if files: self.add_files_to_table(files)

    def add_files_to_table(self, files):
        for f in files:
            row = self.batch_table.rowCount()
            self.batch_table.insertRow(row)
            size = os.path.getsize(f)
            size_str = f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/1024/1024:.1f} MB"
            
            self.batch_table.setItem(row, 0, QTableWidgetItem(os.path.basename(f)))
            self.batch_table.item(row, 0).setToolTip(f)
            self.batch_table.setItem(row, 1, QTableWidgetItem(size_str))
            self.batch_table.setItem(row, 2, QTableWidgetItem("---"))
            self.batch_table.setItem(row, 3, QTableWidgetItem("---"))

    def clear_batch(self):
        self.batch_table.setRowCount(0)

    def process_batch(self):
        for row in range(self.batch_table.rowCount()):
            full_path = self.batch_table.item(row, 0).toolTip()
            try:
                res = self.hash_logic.calculate_file_hash(full_path)
                self.batch_table.setItem(row, 2, QTableWidgetItem(res['md5']))
                self.batch_table.setItem(row, 3, QTableWidgetItem(res['sha256']))
            except:
                self.batch_table.setItem(row, 2, QTableWidgetItem("Error"))

    def export_batch_excel(self):
        if self.batch_table.rowCount() == 0:
            return
        path, _ = QFileDialog.getSaveFileName(self, self.tr('export_batch_results'), "reporte_hashes.xlsx", "Excel (*.xlsx)")
        if path:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                self.tr('col_file'),
                self.tr('col_size'),
                self.tr('col_md5'),
                self.tr('col_sha256'),
                "Ruta completa"
            ])
            for r in range(self.batch_table.rowCount()):
                item_file = self.batch_table.item(r, 0)
                item_size = self.batch_table.item(r, 1)
                item_md5 = self.batch_table.item(r, 2)
                item_sha = self.batch_table.item(r, 3)
                ws.append([
                    item_file.text() if item_file else "",
                    item_size.text() if item_size else "",
                    item_md5.text() if item_md5 else "",
                    item_sha.text() if item_sha else "",
                    item_file.toolTip() if item_file else ""
                ])
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(path)
            QMessageBox.information(self, self.tr('success'), "Reporte exportado correctamente.")


    def setup_duplicate_finder_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        input_layout = QHBoxLayout()
        self.dup_folder_input = QLineEdit()
        self.dup_folder_input.setPlaceholderText(self.tr('select_folder'))
        browse_btn = QPushButton(self.tr('browse'))
        browse_btn.clicked.connect(self.browse_duplicate_folder)
        find_btn = QPushButton(self.tr('find_duplicates'))
        find_btn.clicked.connect(self.find_duplicates)
        
        input_layout.addWidget(self.dup_folder_input)
        input_layout.addWidget(browse_btn)
        input_layout.addWidget(find_btn)
        layout.addLayout(input_layout)
        
        self.dup_results = QTableWidget(0, 3)
        self.dup_results.setHorizontalHeaderLabels([
            self.tr('col_file'), self.tr('col_size'), self.tr('col_md5')
        ])
        self.dup_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.dup_results)
        
        self.tab_widget.addTab(tab, self.tr('duplicate_tab'))

    def setup_pro_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        info = QLabel(self.tr('audit_report_info'))
        info.setWordWrap(True)
        layout.addWidget(info)
        
        input_layout = QHBoxLayout()
        self.report_folder_input = QLineEdit()
        browse_btn = QPushButton(self.tr('browse'))
        browse_btn.clicked.connect(self.browse_report_folder)
        input_layout.addWidget(self.report_folder_input)
        input_layout.addWidget(browse_btn)
        layout.addLayout(input_layout)
        
        gen_btn = QPushButton(self.tr('generate_report'))
        gen_btn.setFixedSize(220, 40)
        gen_btn.setStyleSheet("font-size: 14px; font-weight: bold; color: white;")
        gen_btn.clicked.connect(self.generate_pro_report)
        layout.addWidget(gen_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        self.report_status = QLabel("")
        layout.addWidget(self.report_status)
        layout.addStretch()
        
        self.tab_widget.addTab(tab, self.tr('reports_tab'))

    def browse_duplicate_folder(self):
        folder = QFileDialog.getExistingDirectory(self, self.tr('select_folder'))
        if folder: self.dup_folder_input.setText(folder)

    def browse_report_folder(self):
        folder = QFileDialog.getExistingDirectory(self, self.tr('select_folder'))
        if folder: self.report_folder_input.setText(folder)

    def find_duplicates(self):
        folder = self.dup_folder_input.text()
        if not folder or not os.path.isdir(folder): return
        
        hashes = {} # hash: [paths]
        self.dup_results.setRowCount(0)
        
        for root, _, files in os.walk(folder):
            for f in files:
                path = os.path.join(root, f)
                try:
                    h = self.hash_logic.calculate_file_hash(path)['md5']
                    if h in hashes: hashes[h].append(path)
                    else: hashes[h] = [path]
                except: continue
        
        duplicates = {k: v for k, v in hashes.items() if len(v) > 1}
        
        if not duplicates:
            QMessageBox.information(self, self.tr('success'), self.tr('no_duplicates'))
            return
            
        for h, paths in duplicates.items():
            for p in paths:
                row = self.dup_results.rowCount()
                self.dup_results.insertRow(row)
                size = os.path.getsize(p)
                self.dup_results.setItem(row, 0, QTableWidgetItem(os.path.basename(p)))
                self.dup_results.item(row, 0).setToolTip(p)
                self.dup_results.setItem(row, 1, QTableWidgetItem(f"{size/1024:.1f} KB"))
                self.dup_results.setItem(row, 2, QTableWidgetItem(h))

    def generate_pro_report(self):
        folder = self.report_folder_input.text()
        if not folder or not os.path.isdir(folder): return
        
        save_path, _ = QFileDialog.getSaveFileName(self, "Guardar Reporte", "auditoria_hash.xlsx", "Excel (*.xlsx);;JSON (*.json)")
        if not save_path: return
        
        results = []
        for root, _, files in os.walk(folder):
            for f in files:
                path = os.path.join(root, f)
                try:
                    h = self.hash_logic.calculate_file_hash(path)
                    results.append({
                        'archivo': f,
                        'ruta': path,
                        'tamaño': os.path.getsize(path),
                        'md5': h['md5'],
                        'sha256': h['sha256']
                    })
                except: continue
        
        if save_path.endswith('.json'):
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=4)
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Archivo", "Tamaño", "MD5", "SHA-256", "Ruta"])
            for r in results:
                ws.append([r['archivo'], r['tamaño'], r['md5'], r['sha256'], r['ruta']])
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(save_path)
        
        QMessageBox.information(self, self.tr('success'), "Reporte de auditoría generado correctamente.")

    def apply_styles(self):
        # Cargar estilos desde estilo.qss y aplicar el bloque correspondiente
        qss_path = os.path.join(os.path.dirname(__file__), "estilo.qss")
        with open(qss_path, "r", encoding="utf-8") as f:
            qss = f.read()
        if self.dark_mode:
            # Extraer bloque de modo oscuro
            start = qss.find("/*dark*/")
            end = qss.find("/*light*/")
            dark_qss = qss[start+len("/*dark*/"):end].strip() if start != -1 and end != -1 else ""
            self.setStyleSheet(dark_qss)
        else:
            # Extraer bloque de modo claro
            start = qss.find("/*light*/")
            light_qss = qss[start+len("/*light*/"):].strip() if start != -1 else ""
            self.setStyleSheet(light_qss)

    def toggle_dark_mode(self):
        self.dark_mode = not self.dark_mode
        self.toggle_mode_btn.setText("🌙" if not self.dark_mode else "🌞")
        self.apply_styles()

    def toggle_language(self):
        self.language = 'en' if self.language == 'es' else 'es'
        # Guardar historial actual
        curr_text_hist = self.text_history
        curr_file_hist = self.file_history
        
        # Reiniciar UI
        self.setup_ui()
        
        # Restaurar historial
        self.text_history = curr_text_hist
        self.file_history = curr_file_hist
        self.update_text_history()
        self.update_file_history()

    def calculate_text_hash(self):
        text = self.text_input.text()
        if not text:
            QMessageBox.warning(self, self.tr('warning'), "Por favor ingrese algún texto.")
            return
        
        try:
            results = self.hash_logic.calculate_hashes(text)
            for algo, res in results.items():
                if algo in self.text_results:
                    self.text_results[algo].setText(res)
            
            self.text_history.append({'texto': text, 'hashes': results})
            self.update_text_history()
        except Exception as e:
            QMessageBox.critical(self, self.tr('error'), str(e))

    def calculate_file_hash(self):
        file_path = self.file_input.text()
        if not file_path or not os.path.isfile(file_path):
            QMessageBox.warning(self, self.tr('warning'), "Seleccione un archivo válido.")
            return
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        self.thread = HashCalculatorThread(file_path)
        self.thread.progress.connect(self.progress_bar.setValue)
        self.thread.finished.connect(self.handle_file_hash_result)
        self.thread.error.connect(self.handle_error)
        self.thread.start()

    def handle_file_hash_result(self, results):
        self.progress_bar.setVisible(False)
        for algo, res in results.items():
            if algo in self.file_results:
                self.file_results[algo].setText(res)
        
        self.file_history.append({'archivo': self.file_input.text(), 'hashes': results})
        self.update_file_history()
        QMessageBox.information(self, self.tr('success'), "Hashes calculados exitosamente.")

    def handle_error(self, err_msg):
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, self.tr('error'), err_msg)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, self.tr('select_file'))
        if file_path:
            self.file_input.setText(file_path)

    def browse_verify_file(self, mode):
        file_path, _ = QFileDialog.getOpenFileName(self, self.tr('select_file'))
        if file_path:
            if mode == 0: self.verify_file_input.setText(file_path)
            elif mode == 1: self.compare_file1_input.setText(file_path)
            elif mode == 2: self.compare_file2_input.setText(file_path)

    def browse_folder(self, num):
        folder = QFileDialog.getExistingDirectory(self, self.tr('select_folder'))
        if folder:
            if num == 1: self.folder1_input.setText(folder)
            else: self.folder2_input.setText(folder)

    def verify_hash(self):
        file_path = self.verify_file_input.text()
        exp_hash = self.verify_hash_input.text().strip()
        
        if not file_path or not os.path.isfile(file_path) or not exp_hash:
            QMessageBox.warning(self, self.tr('warning'), "Complete los campos correctamente.")
            return
        
        try:
            res = self.hash_logic.verify_file_integrity(file_path, exp_hash)
            if res['match']:
                self.verify_status_label.setText("✔️ VERIFICACIÓN EXITOSA")
                self.verify_status_label.setStyleSheet("color: green; font-size: 16px; font-weight: bold;")
                self.verify_details_label.setText(f"Algoritmo: {res['algorithm'].upper()}\nHash: {res['calculated_hash']}")
            else:
                self.verify_status_label.setText("❌ VERIFICACIÓN FALLIDA")
                self.verify_status_label.setStyleSheet("color: red; font-size: 16px; font-weight: bold;")
                self.verify_details_label.setText(f"El hash no coincide con ningún algoritmo soportado.\nCalculado (SHA256): {res['calculated_hash']}")
        except Exception as e:
            self.handle_error(str(e))

    def compare_files(self):
        f1 = self.compare_file1_input.text()
        f2 = self.compare_file2_input.text()
        if not f1 or not f2 or not os.path.isfile(f1) or not os.path.isfile(f2):
            QMessageBox.warning(self, self.tr('warning'), "Seleccione dos archivos válidos.")
            return
        
        try:
            res = self.hash_logic.compare_files(f1, f2)
            if res['match']:
                self.verify_status_label.setText("✔️ ARCHIVOS IDÉNTICOS")
                self.verify_status_label.setStyleSheet("color: green; font-size: 16px; font-weight: bold;")
            else:
                self.verify_status_label.setText("❌ ARCHIVOS DIFERENTES")
                self.verify_status_label.setStyleSheet("color: red; font-size: 16px; font-weight: bold;")
            
            details = "Comparación por algoritmo:\n"
            for algo, comp in res['comparisons'].items():
                status = "✔️" if comp['match'] else "❌"
                details += f"{algo.upper()}: {status}\n"
            self.verify_details_label.setText(details)
        except Exception as e:
            self.handle_error(str(e))

    def compare_folders(self):
        dir1 = self.folder1_input.text()
        dir2 = self.folder2_input.text()
        if not dir1 or not dir2 or not os.path.isdir(dir1) or not os.path.isdir(dir2):
            QMessageBox.warning(self, self.tr('warning'), "Seleccione dos carpetas válidas.")
            return
        
        # Lógica simplificada de comparación de carpetas
        files1 = {f: os.path.join(dir1, f) for f in os.listdir(dir1) if os.path.isfile(os.path.join(dir1, f))}
        files2 = {f: os.path.join(dir2, f) for f in os.listdir(dir2) if os.path.isfile(os.path.join(dir2, f))}
        
        all_files = set(files1.keys()) | set(files2.keys())
        results = ""
        
        for f in sorted(all_files):
            if f in files1 and f in files2:
                # Podríamos calcular hashes aquí, pero para rapidez solo comparamos tamaño
                s1 = os.path.getsize(files1[f])
                s2 = os.path.getsize(files2[f])
                if s1 == s2: results += f"[OK] {f} (Mismo tamaño)\n"
                else: results += f"[DIFF] {f} (Diferente tamaño)\n"
            elif f in files1: results += f"[ONLY 1] {f}\n"
            else: results += f"[ONLY 2] {f}\n"
            
        self.folder_compare_results.setText(results)

    def verify_hashes_from_file(self):
        json_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar archivo JSON de hashes", "", "JSON (*.json)")
        if not json_path: return
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Formato esperado: list of {'archivo': path, 'hash': val}
            results = "Resultados de la verificación masiva:\n\n"
            for item in data:
                path = item.get('archivo')
                exp = item.get('hash')
                if path and exp and os.path.isfile(path):
                    res = self.hash_logic.verify_file_integrity(path, exp)
                    status = "✔️ OK" if res['match'] else "❌ ERROR"
                    results += f"{status} - {os.path.basename(path)}\n"
                else:
                    results += f"⚠️ SALTADO - {path} (No encontrado)\n"
            
            self.auto_verify_results.setText(results)
        except Exception as e:
            self.handle_error(str(e))

    def update_text_history(self):
        if self.text_history:
            last = self.text_history[-1]['texto']
            preview = last[:20] + "..." if len(last) > 20 else last
            self.text_history_last.setText(f"{self.tr('last_entry')} {preview}")

    def update_file_history(self):
        if self.file_history:
            last = os.path.basename(self.file_history[-1]['archivo'])
            self.file_history_last.setText(f"{self.tr('last_entry')} {last}")

    def copy_to_clipboard(self, text):
        if text:
            QApplication.clipboard().setText(text)

    def export_text_history(self):
        if not self.text_history:
            return
        path, _ = QFileDialog.getSaveFileName(self, self.tr('export_text_history'), "historial_texto.xlsx", "Excel (*.xlsx)")
        if path:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            # Encabezados dinámicos según los algoritmos presentes
            algos = set()
            for entry in self.text_history:
                algos.update(entry['hashes'].keys())
            algos = sorted(list(algos))
            ws.append(["Texto"] + [a.upper() for a in algos])
            for entry in self.text_history:
                row = [entry['texto']]
                for a in algos:
                    row.append(entry['hashes'].get(a, ""))
                ws.append(row)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(path)
            QMessageBox.information(self, self.tr('success'), "Historial exportado.")

    def export_file_history(self):
        if not self.file_history:
            return
        path, _ = QFileDialog.getSaveFileName(self, self.tr('export_file_history'), "historial_archivos.xlsx", "Excel (*.xlsx)")
        if path:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            algos = set()
            for entry in self.file_history:
                algos.update(entry['hashes'].keys())
            algos = sorted(list(algos))
            ws.append(["Archivo"] + [a.upper() for a in algos])
            for entry in self.file_history:
                row = [entry['archivo']]
                for a in algos:
                    row.append(entry['hashes'].get(a, ""))
                ws.append(row)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(path)
            QMessageBox.information(self, self.tr('success'), "Historial exportado.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = HashGUI()
    window.show()
    sys.exit(app.exec())